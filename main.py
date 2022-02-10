from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import *
import os

workbook = load_workbook('study.xlsx')
sheet = workbook["study_sh"]


def find_cell(num: str):
    for i in sheet.iter_cols():
        if num in str(i[0].value):
            return i[0].column


def write_times(who: str, col):
    for n in sheet.iter_rows():
        row = n[0]
        if who == row.value:
            sheet.cell(row.row,col).value = "√"


if __name__ == "__main__":
    which_periods = input("请输入需要统计青年大学习的期数（如：10 或 10期）")
    colmd = find_cell(which_periods)  # 竖列数
    n = 1
    with open("Finished.txt","r",encoding="utf-8")as nameList:
        getName = nameList.read().split("hh")
    getName.pop()  #移除列表末尾元素
    
    for i in getName:
        write_times(i, colmd)
        print(f"已记录第{n}个，名字为：{i}")
        n += 1
    
    workbook.save("study.xlsx")
    os.remove("Finished.txt")